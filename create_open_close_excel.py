from openpyxl import Workbook, load_workbook
import json
import datetime
import finnhub

def extract_names_from_json(json_file_path):
    with open(json_file_path, "r") as f:
        json_data = json.load(f)

    names = []
    for entry in json_data:
        names.append(entry["Symbol"])

    return names


json_file_path = "s&p500.json"
# names = extract_names_from_json(json_file_path)

names = ["SPY"]
finnhub_client = finnhub.Client(api_key="cl056c1r01qhjei33va0cl056c1r01qhjei33vag")

end_utc = 1699304400 #datetime.datetime.utcnow()
start_utc = 846763200

print(datetime.datetime.fromtimestamp(end_utc))

# start_utc = end_utc - datetime.timedelta(days=1)

print(datetime.datetime.fromtimestamp(start_utc))

max_time_stamps = 0
# Get the stock candles for AAPL for the past minute
# for symbol in names:
#     candles = finnhub_client.stock_candles(symbol, 'D', start_utc, end_utc)
#     # Print the candles
#     max_time_stamps = max(max_time_stamps, len(candles['t']))
#     if max_time_stamps == 6824:
#         print(symbol)

# print(max_time_stamps)

# wb = Workbook()
# ws = wb.create_sheet()

# ws.cell(row=1, column=1).value = "Date"
# # ws.cell(row=1, column=2).value = "Close"

# candles = finnhub_client.stock_candles('SPY', 'D', start_utc, end_utc)

# for row in range(2, len(candles['c']) + 2):
#     date_object = datetime.datetime.fromtimestamp(candles["t"][row - 2])
#     formatted_date = date_object.strftime('%Y-%m-%d')  # Format date without time
#     ws.cell(row=row, column=1).value = formatted_date
#     # ws.cell(row=row, column=2).value = candles["c"][row - 2]

# wb.save("s&p500.xlsx")

# Load the Close-1996.xlsx file
wb = load_workbook('s&p500.xlsx')
ws = wb.active  # Assume the data is in the active sheet

column = 2
# Loop through 500 stocks
for stock_symbol in names:  # Replace with your list of 500 stocks
    column = column + 1
    ws.cell(row=1, column=column).value = stock_symbol
    # Get stock candles for the specified dates
    candles = finnhub_client.stock_candles(stock_symbol, 'D', start_utc, end_utc)
    print("Symbol: ", stock_symbol)
    print("**********************STOCK ", column)

    latest_row = 1

    # Iterate through the stock candles
    for i in range(len(candles['c'])):
        # Convert the timestamp to a datetime object
        date_object = datetime.datetime.fromtimestamp(candles['t'][i])

        # Check if the date matches the dates in the .xlsx file
        row_num = latest_row  # Start searching from the latest row since time-ordered
        date_found = False
        while True:
            if ws.cell(row=row_num, column=1).value == date_object.strftime('%Y-%m-%d'):
                # print(ws.cell(row=row_num, column=1).value)
                date_found = True
                latest_row = row_num
                break
            else:
                row_num += 1
                if ws.cell(row=row_num, column=1).value is None:  # Reached the end of the data
                    break

        # If the date matches, insert the stock data
        if date_found:
            # Insert the closing price into the corresponding column
            ws.cell(row=row_num, column=column).value = candles['c'][i]

    # Save the updated Excel file
    wb.save('s&p500.xlsx')